import asyncio
import json
import multiprocessing
import os
import asyncpg
from pathlib import Path

import openpyxl


def process_task(task_id: str, item: {}) -> ():
    try:
        # такой код будет работать вне зависимости от ОС
        path_to_excel = Path(os.getcwd(), "files_to_send", f"{task_id}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"].value = "Data from client:"
        ws["B1"].value = item["sample_data"]
        wb.save(str(path_to_excel))
        wb.close()
        status = "Done successfully"
    except Exception as e:
        status = "Done with error"
        print(e)
    print(status)
    return status


def process_wrapper(task_id_t, data_t, queue_t):
    result = process_task(task_id_t, data_t)
    queue_t.put(obj=result)


async def worker(name, postgres_pool):
    while True:
        query_str = "WITH next_task as " \
                    "(select id from vkr_schema.documentrequests " \
                    "where status = 'created' limit 1 for update skip locked )" \
                    "update vkr_schema.documentrequests set status = 'processing'" \
                    "from next_task where vkr_schema.documentrequests.id=next_task.id " \
                    "returning vkr_schema.documentrequests.id, vkr_schema.documentrequests.document_type, " \
                    "vkr_schema.documentrequests.data;"
        async with postgres_pool.acquire() as conn:
            data = await conn.fetchrow(query_str)
        if data is None:
            continue
        queue = multiprocessing.Queue()
        process = multiprocessing.Process(target=process_wrapper, args=(data['id'], json.loads(data['data']), queue))
        process.start()
        process.join()
        result = queue.get()
        await asyncio.sleep(10)
        status = result
        query_str = "UPDATE vkr_schema.documentrequests SET status = $1 where id = $2"
        async with postgres_pool.acquire() as conn:
            await conn.fetchrow(query_str, status, data['id'])
        print(f"Worker {name} completed task {data['id']}")
        del queue
        await asyncio.sleep(1)


async def main():
    postgres_pool = await asyncpg.create_pool(
        host="192.168.3.16",
        port="5432",
        database="vkr",
        user="worker",
        password="gjhnatkm"
    )
    workers = [asyncio.create_task(worker(f'worker{i}', postgres_pool)) for i in range(8)]
    await asyncio.gather(*workers)


if __name__ == '__main__':
    asyncio.run(main())
